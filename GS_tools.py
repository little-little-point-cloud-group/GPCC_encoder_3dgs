import os
from pathlib import Path
import filecmp
from tqdm import tqdm
from example.gs_quantize import quantize_3dg, dequantize_3dg
from example.gs_read_write import writePreprossConfig, readPreprossConfig, read3DG_ply, write3DG_ply
import subprocess
import numpy as np
import re
import openpyxl
import shutil
from my_tools import File
import multiprocessing


# 测试分支，注意：此处请单选
branch_selected = (
    # "octree-predlift",
    # "octree-predlift-inter",
     "octree-raht",
    # "octree-raht-inter",
    # "predgeom-predlift",
    # "predgeom-predlift-inter",
    # "predgeom-raht",
    # "predgeom-raht-inter",
)
# 测试条件，注意：此处请单选
condition_selected = {
     "C1": "lossless-geom-lossy-attrs",
    #"C2": "lossy-geom-lossy-attrs",
    # "CW": "lossless-geom-lossless-attrs",
    # "CY": "lossless-geom-nearlossless-attrs",
}

# 点云类别，目前这个变量没有使用
class_selected = (
    #"ManWithFruit",
    "m71763_breakfast_stable",
    #"m71763_cinema_stable",
    #"m71763_bartender_stable",
)

#1F-geom测试条件没有半track
tracks=(
    "track",
    #"partially-track",
)

# 待测试的编解码器,str(tmc3_selected)+'_tmc3.exe'为对应文件
tmc3_selected = 1


# 输入文件路径
template_excel = "ctc/empty.xlsm"  # 带宏的 Excel 模板
output_excel="lift__vs__raht.xlsm"
thread_num_limit=8                     #编码、计算失真的进程数


PCC_sequence=r'D:\pcc_sequence\MPEG_3DGS'

#是否保存图像
save_iamge=0

#保存视角，单帧此参数无效
onlyViewpoint=1


# 渲染图像的宽、高、视角数
seq_information = {
    "ManWithFruit": [3840, 2160, 24],
    "m71763_breakfast_stable": [1920, 1080, 15],
    "m71763_cinema_stable": [1920, 1080, 21],
    "m71763_bartender_stable": [1920, 1080, 21],
}

def pre_process(output,class_selecte,track,frame):
    # Modify these locations to reflect your machine
    pointCloud=PCC_sequence+"/"+class_selecte+"/"+track+"/"+frame+".ply"
    file_raw = Path(pointCloud)  # input: the raw model frame in INRIA format
    file_config = Path(output + "/" + "quantized.json")  # output: json file containing the informarion necessary to inverse the quantization
    file_quantized = Path(output + "/" + "quantized.ply")  # output: PLY file with quantized 3DG attributes

    # Modify these to the desired quantization parameters
    bits_pos = 18
    bits_sh = 12
    bits_opacity = 12
    bits_scale = 12
    bits_rot = 12

    limits_pos = [[0, 0, 0], 256]
    limits_sh = [-4, 4]
    limits_opacity = [-7, 18]

    limits_scale = [-26, 4]
    #limits_scale=np.exp(limits_scale)
    limits_rot = [-1, 1]

    bits = [bits_pos, bits_sh, bits_opacity, bits_scale, bits_rot]
    limits = [limits_pos, limits_sh, limits_opacity, limits_scale, limits_rot]

    # Quantization

    pos, sh, opacity, scale, rot = read3DG_ply(file_raw, tqdm)

    for k in range(3):
        limits[0][0][k] = pos[:, k].min()
    writePreprossConfig(file_config, bits, limits)


    q_pos, q_sh, q_opacity, q_scale, q_rot = quantize_3dg(bits, limits, pos, sh, opacity, scale, rot, tqdm)
    #q_sh[:,1:,:]=2048

    write3DG_ply(q_pos, q_sh, q_opacity, q_scale, q_rot, False, file_quantized, tqdm)

def post_process(output):
    file_decoded = Path(output + "/decoder.ply")  # input: the PLY file of the decoded frame
    file_config = Path(output + "/quantized.json")  # input: json file containing the informarion necessary to inverse the quantization
    file_dequantized = Path(output + "/dequantized.ply")  # output: PLY file of the dequantized decoded frame

    # Dequantization

    q_pos, q_sh, q_opacity, q_scale, q_rot = read3DG_ply(file_decoded, tqdm)
    bits, limits = readPreprossConfig(file_config)


    r_pos, r_sh, r_opacity, r_scale, r_rot = dequantize_3dg(bits, limits, q_pos, q_sh, q_opacity, q_scale, q_rot,
                                                            tqdm)


    write3DG_ply(r_pos, r_sh, r_opacity, r_scale, r_rot, True, file_dequantized, tqdm)

def encoder(output,rate_point,exe,tmc):
    if not os.path.exists(exe):
        print("无启动文件")
    frame=output.split("/")[-1]

    os.makedirs(str(Path(output).parent)+"/txt", exist_ok=True)
    _file =str(Path(output).parent)+"/txt/"+frame+"__Bitbream__encoder.txt"
    with open(_file, "w") as f:
        main=exe
        condition_selecte = condition_selected[list(condition_selected.keys())[0]]
        cfg_path = os.getcwd() + "/cfg" + str(tmc) + "/" + branch_selected[0] + "/" + condition_selecte + "/" + rate_point
        para_encfg = "-c " + cfg_path + "/encoder.cfg"
        para_decfg = "-c " + cfg_path + "/decoder.cfg"

        para2 = "--uncompressedDataPath=" + output + "/" + "quantized.ply"
        para3 = "--compressedStreamPath=" + output + "/" + "compress.bin"
        para4 = "--reconstructedDataPath=" + output + "/" + "encoder.ply"
        para = "%s %s %s %s %s" % (main, para_encfg, para2, para3,para4)
        r = subprocess.run(para, capture_output=True, text=True)
        print(r.stdout, file=f)
        if not os.path.exists(output + "/" + "encoder.ply"):
            print("编码端重建失败")
            print("运行配置为: " + para)

    _file = str(Path(output).parent)+"/txt/"+frame+ "__Bitbream__decoder.txt"
    with open(_file, "w") as f:
        para4 = "--reconstructedDataPath=" + output + "/" + "decoder.ply"
        para = "%s %s %s %s" % (main, para_decfg, para3, para4)
        r = subprocess.run(para, capture_output=True, text=True)
        print(r.stdout, file=f)
        if not os.path.exists(output + "/" + "decoder.ply"):
            print("解码端重建失败")
            print("运行配置为: " + para)
            print(r.stdout)
    if not filecmp.cmp(output + "/" + "encoder.ply", output + "/" + "decoder.ply", shallow=False):
        print("编解码不匹配")

def cam_to_ply(ply,camDIR,exe,output):
    para1="--input="+ply
    para2=para3=""

    if os.path.exists(camDIR+"/cameras.txt"):
        para2="--camera="+camDIR+"/cameras.txt"
    elif os.path.exists(camDIR+"/cameras.bin"):
        para2 = "--camera=" + camDIR + "/cameras.bin"
    else:
        print("没有camera文件")

    if os.path.exists(camDIR+"/images.txt"):
        para3="--image="+camDIR+"/images.txt"
    elif os.path.exists(camDIR+"/images.bin"):
        para3 = "--image=" + camDIR + "/images.bin"
    else:
        print("没有image文件")

    para4="--output="+output
    para5="--verbose=1"
    para = "%s %s %s %s %s %s" % (exe, para1, para2, para3, para4, para5)
    r = subprocess.run(para, capture_output=True, text=True)

def metrics(exe,src,dec,frame_start,frame_num,width,hight,num_view):

    para1="-a "+src
    para2="-b "+dec
    para3="--width="+str(width)+" --height="+str(hight)

    para4="-i "+str(frame_start)+" -f "+str(frame_num)

    if frame_num==1:
        para5 = ("--useCameraPosition=1"
                 f" -s {save_iamge}"
                 f" -n {num_view}")
    else:
        para5=("--useCameraPosition=1"
           f" -s {save_iamge}"
           f" --onlyViewpoint={onlyViewpoint}")

    para = "%s %s %s %s %s %s" % (exe, para1, para2, para3, para4, para5)
    r = subprocess.run(para, capture_output=True, text=True)

    metric_path = str(Path(dec).parent.parent) + "/__metrics.txt"
    with open(metric_path, "w") as f:
        print(r.stdout,file=f)

class Gaussian:
    def __init__(self,PCC_sequence=PCC_sequence):

        self.frames = range(0, 2)  # 起始帧，终止帧,指stable的起始帧与结束帧
        self.fruitFrame = range(51, 52)

        self.anchor_name="lift"
        self.test_name="raht"
# ======================================
# 下面变量代码为测试的运行文件

        self.tmc13=str(tmc3_selected)+'_tmc3.exe'

# ======================================
# 下面变量代码短期不需要修改

        self.PCC_sequence=PCC_sequence
        self.cameraPosition="example/cameraPosition.exe"
        self.mpeg_gsc_metrics = "example/mpeg-gsc-metrics.exe"
        self.rate_points=["r01","r02","r03","r04",]
        #self.rate_points=["r01",]



        self.anchor_columns = {
            "PSNR-RGB": "F",  # PSNR-RGB 列
            "PSNR-YCbCr": "G",  # PSNR-YCbCr 列
            "SSIM-YCbCr": "H",  # SSIM-YCbCr 列
            #"IVSSIM": "I",  # IVSSIM 列
            #"LPIPS": "J"  # LPIPS 列
        }

        self.test_columns = {
            "PSNR-RGB": "AD",  # PSNR-RGB 列
            "PSNR-YCbCr": "AE",  # PSNR-YCbCr 列
            "SSIM-YCbCr": "AF",  # SSIM-YCbCr 列
            #"IVSSIM": "AA",  # IVSSIM 列
            #"LPIPS": "AB"  # LPIPS 列
        }

        self.PSNR_start_row = 33

        self.anchor_Bitstream_columns={
        "position":"E",
        "sh0":"F",
        "sh1":"G",
        "sh2":"H",
        "sh3":"I",
        "rotation":"J",
        "scaling":"K",
        "opacity":"L",
        "metadata":"M",
        "T_Enc":"P",
        "T_Dec":"Q",
        "G_Enc": "R",
        "G_Dec": "S",
        "A_Enc": "T",
        "A_Dec": "U",
        "maxRSS_Enc": "V",
        "maxRSS_Dec": "W",
        }

        self.test_Bitstream_columns={
        "position":"AC",
        "sh0":"AD",
        "sh1":"AE",
        "sh2":"AF",
        "sh3":"AG",
        "rotation":"AH",
        "scaling":"AI",
        "opacity":"AJ",
        "metadata":"AK",
        "T_Enc":"AN",
        "T_Dec":"AO",
        "G_Enc": "AP",
        "G_Dec": "AQ",
        "A_Enc": "AR",
        "A_Dec": "AS",
        "maxRSS_Enc": "AT",
        "maxRSS_Dec": "AU",
        }


        # 复制工作表
        self.tracks_name = {
            "track": "track",
            "partially-track": "semitrack",
        }

        self.Bitstream_start_row = 16
# ======================================
#下面变量实时更新
        self.condition_selecte = None
        self.class_selecte = class_selected[0]
        self.metrics_path = dict()
        self.tmc=tmc3_selected
        # 加载带宏的模板文件
        self.wb =openpyxl.load_workbook(template_excel, keep_vba=True,read_only=False)

        def set_name():
            self.wb["Summary (auto)"].cell(row=3, column=3).value = self.anchor_name
            self.wb["Summary (auto)"].cell(row=4, column=3).value = self.test_name

        self.set_name=set_name

    def run_with_anchor(self):

        self.wb = openpyxl.load_workbook(template_excel, keep_vba=True, read_only=False)
        self.run()
        self.tmc=0
        self.tmc13 = '0_tmc3.exe'
        self.wb = openpyxl.load_workbook(branch_selected[0]+"/"+output_excel, keep_vba=True, read_only=False)
        self.run()

    def run(self):
            print("代码：" + self.tmc13)


        #清空分支
            if os.path.exists(branch_selected[0]):
                shutil.rmtree(branch_selected[0])

        # ======================================
        # 编解码
            thread_pool = multiprocessing.Pool(thread_num_limit)
            for class_selecte in class_selected:
                for track in tracks if class_selecte.find("stable")>=0 else [class_selecte]:
                    for rate_point in self.rate_points:
                        if class_selecte.find("stable")>=0:
                            frames=self.frames
                        else:
                            frames=self.fruitFrame

                        for frame in frames :

                            if class_selecte.find("stable") >= 0:
                                if class_selecte == "m71763_bartender_stable" or class_selecte == "m71763_breakdance_stable":
                                    cameras_name = f"/frame{frame:03d}"
                                else:
                                    cameras_name = ""
                            else:
                                cameras_name = f"/{frame:06d}/sparse/0"

                            cameras_path = PCC_sequence + "/" + class_selecte + "/cameras" + cameras_name
                            #self.sub_run(class_selecte,track,rate_point,frame,self.tmc13,self.tmc,cameras_path,self.cameraPosition)
                            thread_pool.apply_async(self.sub_run, args=(class_selecte,track,rate_point,frame,self.tmc13,self.tmc,cameras_path,self.cameraPosition))

            thread_pool.close()  # 关闭进程池入口，不再接受新进程插入
            thread_pool.join()  # 主进程阻塞，等待进程池中的所有子进程结束，再继续运行主进程

        # ======================================
        # 渲染 计算失真
            thread_pool = multiprocessing.Pool(thread_num_limit)
            for class_selecte in class_selected:
                for track in tracks if class_selecte.find("stable")>=0 else [class_selecte]:
                    for rate_point in self.rate_points:
                        if class_selecte.find("stable") >= 0:
                            frames = self.frames
                        else:
                            frames = self.fruitFrame


                        condition_selecte = condition_selected[list(condition_selected.keys())[0]]
                        DIR=os.getcwd()+"/"+branch_selected[0]+"/"+condition_selecte+"/"+class_selecte+"/"+track+"/"+rate_point

                        thread_pool.apply_async(self.render, args=(frames,DIR,self.mpeg_gsc_metrics,class_selecte))


            thread_pool.close()  # 关闭进程池入口，不再接受新进程插入
            thread_pool.join()  # 主进程阻塞，等待进程池中的所有子进程结束，再继续运行主进程

            self.write_to_excel()      #写入excel

    @staticmethod
    def sub_run(class_selecte,track,rate_point,frame,tmc13,tmc,cameras_path,cameraPosition):
        frame_DIR=frame

        condition_selecte=condition_selected[list(condition_selected.keys())[0]]
        output =  os.getcwd()+"/"+branch_selected[0]+"/"+condition_selecte+"/"+class_selecte+"/"+track+"/"+rate_point+"/"+f"frame{frame:03d}"
        print("正在运行:"+os.path.join(branch_selected[0],condition_selecte,class_selecte,track,rate_point,f"frame{frame:03d}"))
        os.makedirs(output, exist_ok=True)
        frame=f"frame{frame:03d}" if class_selecte.find("stable")>=0 else f"{frame:04d}"

        pre_process(output,class_selecte,track,frame)  # 预处理
        encoder(output,rate_point,tmc13,tmc)  # 编码
        post_process(output)  # 后处理

        pointCloud = PCC_sequence + "/" + class_selecte + "/" + track + "/" + frame + ".ply"

        src_DIR=os.getcwd()+"/"+branch_selected[0]+"/"+condition_selecte+"/"+class_selecte+"/"+track+"/"+rate_point+"/src"
        dec_DIR = os.getcwd()+"/"+branch_selected[0]+"/"+condition_selecte+"/"+class_selecte+"/"+track+"/"+rate_point+"/dec"
        os.makedirs(src_DIR, exist_ok=True)
        os.makedirs(dec_DIR, exist_ok=True)

        cam_to_ply(pointCloud,cameras_path,cameraPosition,src_DIR+"/"+f"frame{frame_DIR:03d}"+".ply")
        cam_to_ply(output + "/dequantized.ply", cameras_path, cameraPosition, dec_DIR + "/" + f"frame{frame_DIR:03d}" + ".ply")

        shutil.rmtree(output)

        print("结束:" + os.path.join(branch_selected[0],condition_selecte,class_selecte,track,rate_point,frame))

    @staticmethod
    def render(frames,DIR,exe,class_selecte):

        src=DIR+"/src/"+"frame"+"%03d" + ".ply"
        dec=DIR+"/dec/"+"frame"+"%03d" + ".ply"
        metrics(exe,src,dec,frames[0],len(frames),seq_information[class_selecte][0],seq_information[class_selecte][1],seq_information[class_selecte][2])


    def write_to_excel(self):
        # ======================================
        # 写入PSNR
        metrics_path=File.get_all_file_from_baseCatalog("__metrics.txt",branch_selected[0])
        metrics_data=[]
        for path in metrics_path:
            metrics_data.append(self.extract_metrics(path))

        self.sub_write_to_excel(metrics_data)


        # ======================================
        # 写入编码比特
        # ======================================
        bitstream_files=File.get_all_file_from_baseCatalog("__Bitbream__encoder.txt",branch_selected[0])
        for bitstream_file in bitstream_files:
            frameId = int(bitstream_file.split("/")[-1].split("frame")[1].split("__")[0])
            class_selecte = bitstream_file.split("/")[-5]

            if class_selecte.find("stable") >= 0:
                frames = self.frames
            else:
                frames = self.fruitFrame

            if frameId in frames:

                parsed_data = self.parse_bitstream(bitstream_file)
                aggregated = self.aggregate_attributes(parsed_data)
                self.sub_bitstream_write_to_excel(aggregated,bitstream_file)

        self.set_name()

        # 保存为新文件（保留宏）
        self.wb.save(branch_selected[0]+"/"+output_excel)
        print(f"数据已成功写入 {output_excel}")
        # 保存为新文件（保留宏）
        os.makedirs("32F-geo", exist_ok=True)
        self.wb.save("1F-geo/"+output_excel)

# ======================================
# 步骤 1：从 metrics.txt 提取数据
# ======================================
    def extract_metrics(self,file_path):
        data=dict()
        PSNR = dict()
        with open(file_path, "r") as f:
            contents = f.readlines()

        for content in contents:
            if content.find("Psnr RGB (avg)")>=0:
                PSNR["PSNR-RGB"]=float(content.split()[4])
            elif content.find("Psnr YUV (wavg)")>=0:
                PSNR["PSNR-YCbCr"]=float(content.split()[4])
            elif content.find("SSIM Avg")>=0:
                PSNR["SSIM-YCbCr"]=float(content.split()[3])

        data[file_path]=PSNR
        return data


# ======================================
# 步骤 2：写入 Excel .xlsm 文件
# ======================================
    def sub_write_to_excel(self,metrics_data):

        columns=self.test_columns if self.tmc else self.anchor_columns

        # 定义数据列的起始位置（根据你的 Excel 模板调整）
        # 示例：假设表头在行1，数据从行2开始
        start_row = self.PSNR_start_row
        # 按图像编号顺序写入数据
        for path in metrics_data:
            key=list(path.keys())[0]
            labels=key.split("/")
            increase_row=int(labels[-2][1:3])-1

            _class = labels[-4].split("_")[1] if labels[-4].find("stable")>=0 else labels[-4]
            track=labels[-3]
            sheet_name = f'{_class}_{self.tracks_name[track]}' if labels[-4].find("stable")>=0 else f'{_class}'

            row = start_row +increase_row# 假设数据按顺序排列
            data = path[key]
            ws=self.wb[sheet_name]
            # 写入指标数据（保留原始精度）
            for key in columns.keys():
                ws[f'{columns[key]}{row}'].value = data[key]


# ======================================
# 步骤 3：提取属性的编码bit
# ======================================
    def parse_bitstream(self,file_path):
        """使用正则表达式解析比特流文件"""
        attributes = dict()
        attributes["e-gtime"]=0
        attributes["e-atime"]=0
        attributes["d-gtime"]=0
        attributes["d-atime"]=0;
        pattern = re.compile(r'^(\w+).*bitstream size (\d+) B \((\d+\.\d+) bpp\)')

        with open(file_path, 'r') as f:
            for line in f:
                match = pattern.match(line.strip())
                if match:
                    attr = match.group(1)  # 属性名
                    size = int(match.group(2))  # 字节数

                    if attr not in attributes:
                        attributes[attr] = 0
                    attributes[attr] +=  size

        total=0
        for key in attributes.keys():
            total+=attributes[key]

        pattern = re.compile(r'^Total bitstream size (\d+) B')
        with open(file_path, 'r') as f:
            for line in f:
                match = pattern.match(line.strip())
                if match:
                    size = int(match.group(1))  # 字节数
                    attributes["Total bitstream size"] = size

        attributes["metadata"] = attributes["Total bitstream size"] - total

        pattern = re.compile(r'^Processing time \(user\): (\d+(?:\.\d+)?) s')
        with open(file_path, 'r') as f:
            for line in f:
                match = pattern.match(line.strip())
                if match:
                    encoder_time = float(match.group(1))  # 字节数
                    attributes["encoder Processing time (user):"] = encoder_time

        with open(file_path, 'r') as f:
            for l in f:
                line=l.split()
                if len(line)<5:
                    continue
                if line[1]=="processing" and line[2]=="time":
                    if line[0]=="positions":
                        attributes["e-gtime"] += float(line[4])
                    else:
                        attributes["e-atime"]+=float(line[4])


        decode_path=file_path.split("__Bitbream__encoder.txt")[0]+"__Bitbream__decoder.txt"

        pattern = re.compile(r'^Processing time \(user\): (\d+(?:\.\d+)?) s')
        with open(decode_path, 'r') as f:
            for line in f:
                match = pattern.match(line.strip())
                if match:
                    decoder_time = float(match.group(1))  # 字节数
                    attributes["decoder Processing time (user):"] = decoder_time

        with open(file_path, 'r') as f:
            for l in f:
                line=l.split()
                if len(line)<5:
                    continue
                if line[1]=="processing" and line[2]=="time":
                    if line[0]=="positions":
                        attributes["d-gtime"] += float(line[4])
                    else:
                        attributes["d-atime"]+=float(line[4])

        return attributes

# ======================================
# 步骤 4：属性聚类
# ======================================
    def aggregate_attributes(self,attrs):
        """聚合属性到指定分类"""
        #第一中定义：我认为的
        sh1=["f_rest_0s","f_rest_1s","f_rest_2s",
             "f_rest_15s","f_rest_16s","f_rest_17s",
             "f_rest_30s","f_rest_31s","f_rest_32s"]
        sh2=["f_rest_3s","f_rest_4s","f_rest_5s","f_rest_6s","f_rest_7s",
             "f_rest_18s","f_rest_19s","f_rest_20s","f_rest_21s","f_rest_22s",
             "f_rest_33s","f_rest_34s","f_rest_35s","f_rest_36s","f_rest_37s"]
        sh3=["f_rest_8s","f_rest_9s","f_rest_10s","f_rest_11s","f_rest_12s","f_rest_13s","f_rest_14s",
             "f_rest_23s","f_rest_24s","f_rest_25s","f_rest_26s","f_rest_27s","f_rest_28s","f_rest_29s",
             "f_rest_38s","f_rest_39s","f_rest_40s","f_rest_41s","f_rest_42s","f_rest_43s","f_rest_44s"]

        #第二种定义：通过结果推理出的
        '''
        sh1=[]
        sh2=[]
        sh3=[]
        for i in range(9):
            sh1.append(f"f_rest_{str(i)}s")
        for i in range(9,24):
            sh2.append(f"f_rest_{str(i)}s")
        for i in range(24,45):
            sh3.append(f"f_rest_{str(i)}s")
        '''

        return {
            "position": sum(attrs[k] for k in attrs.keys() if k.startswith("positions")),
            "sh0": sum(attrs[k] for k in attrs.keys() if k.startswith("f_dc_")),
            "sh1": sum(attrs[k] for k in attrs.keys() if k in sh1),
            "sh2": sum(attrs[k] for k in attrs.keys() if k in sh2),
            "sh3": sum(attrs[k] for k in attrs.keys() if k in sh3),
            "rotation": sum(attrs[k] for k in attrs.keys() if k.startswith("rot_")),
            "scaling": sum(attrs[k] for k in attrs.keys() if k.startswith("scale_")),
            "opacity": sum(attrs[k] for k in attrs.keys() if k.startswith("opacity")),
            "metadata": attrs["metadata"],  # 元数据占位符
            "T_Enc":attrs["encoder Processing time (user):"],
            "T_Dec":attrs["decoder Processing time (user):"],
            "G_Enc": attrs["e-gtime"],
            "G_Dec": attrs["d-gtime"],
            "A_Enc": attrs["e-atime"],
            "A_Dec": attrs["d-atime"],
        }

# ======================================
# 步骤 5：编码比特写入
# ======================================
    def sub_bitstream_write_to_excel(self,data,bitstream_file):
        # 列映射关系
        columns = self.test_Bitstream_columns if self.tmc else self.anchor_Bitstream_columns
        attrs=["position","sh0","sh1", "sh2", "sh3","rotation","scaling","opacity","metadata"]
        # 定义数据列的起始位置（根据你的 Excel 模板调整）
        # 示例：假设表头在行1，数据从行2开始
        start_row = self.Bitstream_start_row
        labels = bitstream_file.split("/")
        increase_row = int(labels[-3][1:3]) - 1

        _class = labels[-5].split("_")[1] if labels[-5].find("stable") >= 0 else labels[-5]
        track = labels[-4]
        sheet_name = f'{_class}_{self.tracks_name[track]}' if labels[-5].find("stable") >= 0 else f'{_class}'

        row = start_row + increase_row  # 假设数据按顺序排列
        for key in data.keys():
            value = data[key]
            ws = self.wb[sheet_name]
            a = ws[f'{columns[key]}{row}'].value
            if a is None:
                ws[f'{columns[key]}{row}'] = 0
                if key in attrs:
                    ws[f'{columns[key]}{row}'].value += data[key]*8/1000
                else:
                    ws[f'{columns[key]}{row}'].value += data[key]
            else:
                if key in attrs:
                    ws[f'{columns[key]}{row}'].value += data[key] * 8 / 1000
                else:
                    ws[f'{columns[key]}{row}'].value += data[key]

    #第一个num表示other中的anchor还是test，第二个同理
    def copy_to_excel(self,other,my,num0,num1):
        other = openpyxl.load_workbook(other, keep_vba=True, read_only=False)
        my = openpyxl.load_workbook(my, keep_vba=True, read_only=False)


        columns0 = self.test_Bitstream_columns if num0 else self.anchor_Bitstream_columns
        columns1 = self.test_Bitstream_columns if num1 else self.anchor_Bitstream_columns


        sheetnames=other.get_sheet_names()
        for i in range(6,len(sheetnames)):
            sheetname=sheetnames[i]
            ws0 = other[sheetname]
            ws1 = my[sheetname]
            start_row = 16
            for increase_row in range(len(self.rate_points)):
                row=start_row+increase_row
                for key in columns1:
                    ws1[f'{columns1[key]}{row}'].value = ws0[f'{columns0[key]}{row}'].value

        columns0 = self.test_columns if num0 else self.anchor_columns
        columns1 = self.test_columns if num1 else self.anchor_columns


        start_row = self.PSNR_start_row
        for i in range(6,len(sheetnames)):
            sheetname=sheetnames[i]
            ws0 = other[sheetname]
            ws1 = my[sheetname]
            for view in range(121):
                for increase_row in range(len(self.rate_points)):
                    row = start_row + 5*view+increase_row# 假设数据按顺序排列
                    for key in columns1:
                        ws1[f'{columns1[key]}{row}'].value = ws0[f'{columns0[key]}{row}'].value



        my.save("1F-geo/" + output_excel)


if __name__ == '__main__':
    pass